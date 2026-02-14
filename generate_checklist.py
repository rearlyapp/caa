#!/usr/bin/env python3
"""
Generate the 'Checklist for Company Formation' Excel template from code.

This makes the template version-controllable and reproducible.

Usage:
    python generate_checklist.py [--output FILENAME]
"""

import argparse
from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ── Template Data ────────────────────────────────────────────────────────────

COMPANY_OBJECTIVES = """1. To carry on the business of wholesale trading, buying, selling, importing, exporting, processing, manufacturing, distributing, storing, transporting and otherwise dealing in all types of agricultural produce, commodities, raw materials, straw, fodder, animal and poultry feed, organic and inorganic products, seeds, fertilizers, pesticides, herbicides, farm machinery and equipment, and all other products and materials related to the agriculture and allied sectors.
2. To carry on the business of commission agents, brokers, clearing and forwarding agents, dealers, distributors, stockists, wholesalers, retailers, franchisees, agents and representatives in respect of all kinds of agricultural produce, commodities and related products.
3. To establish, acquire, maintain and operate processing units, manufacturing plants, cold storage facilities, warehouses, godowns, silos, drying yards, grading units, sorting centres, packaging units and quality testing laboratories for agricultural produce and commodities."""

OTHER_OBJECTIVES = """1. To purchase, take on lease, hire, exchange or otherwise acquire lands, buildings, warehouses, cold storages, godowns, factories, processing units, workshops, offices, plants, machinery, vehicles and equipment required for the business of the Company.
2. To establish, maintain and operate storage facilities, silos, warehouses, cold storage chains, logistics infrastructure and transportation facilities for agricultural raw materials, straw, fodder, animal and poultry feed and other products.
3. To enter into agreements, contracts, arrangements, collaborations, joint ventures, technical collaborations, agency or distributorship agreements with individuals, firms, companies, cooperatives, government bodies or other entities in India or abroad.
4. To appoint agents, brokers, commission agents, clearing and forwarding agents, distributors, franchisees, contractors and other intermediaries for carrying on the business of the Company.
5. To import, export, buy, sell or otherwise deal in raw materials, finished goods, machinery, spare parts and consumables necessary for carrying on the business of the Company.
6. To obtain, apply for, purchase or otherwise acquire and protect licenses, registrations, permits, quotas, approvals, trademarks, copyrights, patents, trade names and other intellectual property rights required for the business.
7. To borrow, raise or secure money by way of loans, overdrafts, cash credit facilities, debentures or other instruments, with or without security, for the purposes of the Company, subject to the provisions of the Companies Act, 2013.
8. To open, operate and close bank accounts, obtain credit facilities, guarantees and letters of credit, and to negotiate and discount bills of exchange, promissory notes and other negotiable instruments.
9. To invest surplus funds of the Company in such manner as may be permitted under applicable laws.
10. To undertake marketing, branding, advertising, promotion, packaging, labeling and quality certification activities for products of the Company.
11. To employ, appoint and remunerate managers, consultants, professionals, technical experts, workers and other personnel for carrying on the business of the Company.
12. To insure the properties, assets, stocks and operations of the Company against all risks.
13. To apply for and obtain government subsidies, incentives, registrations, recognitions and benefits under Central or State Government schemes including MSME, Startup India or agricultural schemes.
14. To undertake research and development, testing, quality control and product improvement activities.
15. To do all such other lawful acts, deeds and things as are incidental or conducive to the attainment of the above main objects."""

# Row definitions: (serial_no_or_section, field_label)
DIRECTOR_FIELDS = [
    ("A", "Information of Directors", True),   # Section header
    ("1",  "NAME as per PAN Card", False),
    ("2",  "Fathers Name", False),
    ("3",  "DOB (DD-MM-YYYY)", False),
    ("4",  "PLACE OF BIRTH", False),
    ("5",  "Nationality", False),
    ("6",  "Resident of India", False),
    ("7",  "OCCUPATION", False),
    ("8",  "EDUCATIONAL QUALIFICATION", False),
    ("9",  "NO. OF SHARES SUBSCRIBED", False),
    ("10", "DURATION OF STAY AT PRESENT ADDRESS (Years & Month)", False),
    ("11", "EMAIL ID", False),
    ("12", "MOBILE NO.", False),
    ("13", "PAN Number", False),
    ("14", "Identity Proof (DL, AADHAR, VOTER ID, PASSPORT) (NO.)", False),
    ("15", "Residential Proof with Address\n(Telephone bill/MOBILE BILL, ELECTRICITY BILL, BANK STATEMENT)", False),
    ("16", "DIN number of directors, if available", False),
    ("17", "ALL Directors Photo (JPEG/JPG)", False),
    ("18", "ALL Directors Signature (PDF)", False),
    ("19", "Form DIR-2 (ALL Directors Merge)", False),
]

COMPANY_FIELDS = [
    ("B", "Information of Company", True),    # Section header
    ("20", "ELECTRICITY BILL/ANY UTILITY BILL CURRENT (Not older than 2 months)", False),
    ("21", "Latitude", False),
    ("22", "Longitude", False),
    ("23", "NOC from Landlord", False),
    ("24", "Office emailid", False),
    ("25", "Office mob. No.", False),
    ("26", "Total Authorized share capital", False),
    ("27", "Total Paid-up share capital", False),
    ("28", "Objectives of the Company", False),
    ("29", "Other Objectives", False),
]

PROFESSIONAL_FIELDS = [
    ("C", "Information of Professional", True),  # Section header
    ("30", "Name of the CA/CS/Advocate", False),
    ("31", "Membership No.", False),
    ("32", "Address ", False),
]


# ── Styles ───────────────────────────────────────────────────────────────────

TITLE_FONT = Font(name="Calibri", size=14, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
NORMAL_FONT = Font(name="Calibri", size=10)

HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ── Generator ────────────────────────────────────────────────────────────────

def generate_checklist(output_path: str):
    """Generate the Company Formation checklist Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Automation Excel"

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35

    row = 1

    # ── Title Row ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    title_cell = ws.cell(row=row, column=1, value="CHECKLIST FOR COMPANY FORMATION")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN
    row += 1

    # ── Header Row ───────────────────────────────────────────────────────
    headers = ["S. NO.", "DOCUMENTS", "Director 1", "Director 2"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    # ── Write sections ───────────────────────────────────────────────────
    all_fields = DIRECTOR_FIELDS + COMPANY_FIELDS + PROFESSIONAL_FIELDS

    for serial, label, is_section in all_fields:
        cell_a = ws.cell(row=row, column=1, value=serial)
        cell_b = ws.cell(row=row, column=2, value=label)

        if is_section:
            # Section header styling
            cell_a.font = SECTION_FONT
            cell_b.font = SECTION_FONT
            cell_a.fill = SECTION_FILL
            cell_b.fill = SECTION_FILL
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = SECTION_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER
        else:
            cell_a.font = NORMAL_FONT
            cell_a.alignment = CENTER_ALIGN
            cell_b.font = NORMAL_FONT
            cell_b.alignment = LEFT_ALIGN

            # Pre-fill objectives if applicable
            if label == "Objectives of the Company":
                ws.cell(row=row, column=3, value=COMPANY_OBJECTIVES).font = NORMAL_FONT
                ws.cell(row=row, column=3).alignment = TOP_LEFT_ALIGN
                ws.row_dimensions[row].height = 80
            elif label == "Other Objectives":
                ws.cell(row=row, column=3, value=OTHER_OBJECTIVES).font = NORMAL_FONT
                ws.cell(row=row, column=3).alignment = TOP_LEFT_ALIGN
                ws.row_dimensions[row].height = 200

            # Apply borders to all 4 columns
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # ── Auto-height for wrapped text rows ────────────────────────────────
    # Row 15 (Residential Proof) gets extra height for multiline label
    for r in range(1, row):
        val = ws.cell(row=r, column=2).value
        if val and "\n" in str(val):
            ws.row_dimensions[r].height = 40

    wb.save(output_path)
    print(f"✅ Checklist template generated: {output_path}")
    print(f"   Rows: {row - 1}, Columns: A–D")
    print(f"   Sections: Directors (rows 3–22), Company (23–33), Professional (34–37)")


def main():
    parser = argparse.ArgumentParser(description="Generate Company Formation checklist template")
    today = date.today().strftime("%d %m %Y")
    default_name = f"Checklist for Company Formation dated {today}.xlsx"
    parser.add_argument("--output", "-o", type=str, default=default_name,
                        help=f"Output filename (default: {default_name})")
    args = parser.parse_args()
    generate_checklist(args.output)


if __name__ == "__main__":
    main()
