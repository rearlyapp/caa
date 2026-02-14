#!/usr/bin/env python3
"""
Fast PAN + Aadhaar extraction using PaddleOCR (traditional OCR) + regex parsing.
Runs in <5 seconds on CPU. No GPU needed.

Usage:
    python extract.py --pan pan.jpg --aadhaar aadhar.jpg --director 1
"""

import argparse
import json
import re
import sys
import time
from pathlib import Path

from paddleocr import PaddleOCR


# ── OCR Engine ───────────────────────────────────────────────────────────────

def get_ocr():
    """Initialize PaddleOCR (lightweight, CPU-friendly)."""
    return PaddleOCR(use_textline_orientation=True, lang='en')


def ocr_image(ocr, image_path: str) -> list[str]:
    """Run OCR on an image and return all detected text lines."""
    result = ocr.ocr(image_path, cls=True)
    lines = []
    if result and result[0]:
        for line in result[0]:
            text = line[1][0]  # (bbox, (text, confidence))
            lines.append(text)
    return lines


# ── PAN Card Parser ──────────────────────────────────────────────────────────

def parse_pan(lines: list[str]) -> dict:
    """Extract PAN card fields from OCR text lines."""
    data = {
        "name": "",
        "fathers_name": "",
        "date_of_birth": "",
        "pan_number": "",
    }

    # PAN number: 5 letters + 4 digits + 1 letter (e.g., BQJPK6347Q)
    pan_pattern = re.compile(r'\b[A-Z]{5}[0-9]{4}[A-Z]\b')

    # Date pattern: DD/MM/YYYY or DD-MM-YYYY
    date_pattern = re.compile(r'\b(\d{2}[/-]\d{2}[/-]\d{4})\b')

    # Skip lines (headers, labels, noise)
    skip_words = [
        'income tax', 'department', 'govt', 'india', 'permanent',
        'account', 'number', 'signature', 'आयकर', 'विभाग', 'भारत',
        'सरकार', 'card', 'lost', 'found', 'please', 'inform',
        'return', 'nsdl', 'floor', 'sapphire', 'chambers', 'near',
        'baner', 'telephone', 'exchange', 'pune', 'tel:', 'fax:',
        'e-mail', 'tininfo', 'this', 'someone'
    ]

    name_candidates = []

    for line in lines:
        text = line.strip()
        if not text:
            continue

        # Extract PAN number
        pan_match = pan_pattern.search(text)
        if pan_match:
            data["pan_number"] = pan_match.group()
            continue

        # Extract date
        date_match = date_pattern.search(text)
        if date_match:
            data["date_of_birth"] = date_match.group().replace('-', '/')
            continue

        # Skip header/noise lines
        lower = text.lower()
        if any(w in lower for w in skip_words):
            continue

        # Skip very short or numeric-only lines
        if len(text) < 3 or text.replace(' ', '').isdigit():
            continue

        # Remaining text lines are likely names
        # PAN cards: line 1 = cardholder name, line 2 = father's name
        if text[0].isupper() and re.match(r'^[A-Z\s\.]+$', text):
            name_candidates.append(text)

    # Assign names: first is cardholder, second is father
    if len(name_candidates) >= 2:
        data["name"] = name_candidates[0].strip()
        data["fathers_name"] = name_candidates[1].strip()
    elif len(name_candidates) == 1:
        data["name"] = name_candidates[0].strip()

    return data


# ── Aadhaar Card Parser ─────────────────────────────────────────────────────

def parse_aadhaar(lines: list[str]) -> dict:
    """Extract Aadhaar card fields from OCR text lines."""
    data = {
        "name": "",
        "aadhaar_number": "",
        "date_of_birth": "",
        "gender": "",
        "address": "",
    }

    # Aadhaar number: 12 digits (may have spaces: XXXX XXXX XXXX)
    aadhaar_pattern = re.compile(r'\b(\d{4}\s?\d{4}\s?\d{4})\b')

    # Date pattern
    date_pattern = re.compile(r'\b(\d{2}[/-]\d{2}[/-]\d{4})\b')

    # Gender
    gender_pattern = re.compile(r'\b(MALE|FEMALE|male|female|Male|Female)\b', re.IGNORECASE)

    # Skip noise
    skip_words = [
        'unique identification', 'authority', 'government',
        'aadhaar', 'aadhar', 'आधार', 'enrolment', 'download',
        'validity', 'valid', 'proof', 'citizenship', 'establish',
        'identity', 'authenticate', 'electronically', 'generated',
        'information', 'helpful', 'availing', 'services', 'future',
        'your aadhaar', 'எனது', 'அடையாளம்', 'உங்கள்',
        'இந்திய', 'அரசாங்கம்', 'தனிப்பட்ட',
        'www.uidai', 'help@uidai', '1947',
    ]

    # Address keywords
    address_keywords = [
        'street', 'road', 'nagar', 'colony', 'lane', 'floor',
        'block', 'sector', 'district', 'state', 'pin', 'south',
        'north', 'east', 'west', 'near', 'opp', 'post',
        'ram nagar', 's s colony', 'arasaradi', 'madurai',
        'tamil nadu', 'tamil', 'nadu',
    ]

    addresses = []
    names_found = []
    aadhaar_numbers = []

    for line in lines:
        text = line.strip()
        if not text:
            continue

        lower = text.lower()

        # Skip noise lines
        if any(w in lower for w in skip_words):
            continue

        # Extract Aadhaar number
        aadhaar_match = aadhaar_pattern.search(text.replace(' ', '  '))
        # Try on original text too
        digits_only = re.sub(r'[^\d]', '', text)
        if len(digits_only) == 12:
            formatted = f"{digits_only[:4]} {digits_only[4:8]} {digits_only[8:]}"
            if formatted not in aadhaar_numbers:
                aadhaar_numbers.append(formatted)
            continue

        # Extract date
        date_match = date_pattern.search(text)
        if date_match:
            data["date_of_birth"] = date_match.group().replace('-', '/')
            # Check for gender on same line (common: "DOB: 03/11/1988")
            continue

        # Extract gender
        gender_match = gender_pattern.search(text)
        if gender_match:
            data["gender"] = gender_match.group().upper()
            continue

        # Skip Tamil/Hindi script lines (not useful for English extraction)
        if re.match(r'^[\u0B80-\u0BFF\u0900-\u097F\s]+$', text):
            continue

        # Address detection
        if any(kw in lower for kw in address_keywords) or re.search(r'\b\d{6}\b', text):
            addresses.append(text)
            continue

        # Name detection: English uppercase/titlecase text
        if re.match(r'^[A-Za-z\s\.\/]+$', text) and len(text) > 3:
            # Skip "S/O", "D/O", "W/O" prefixed lines but extract father name
            so_match = re.match(r'^[SDW]/O\s*:?\s*(.+)', text, re.IGNORECASE)
            if so_match:
                # This contains father's/husband's name (not needed for Aadhaar output but useful)
                continue
            names_found.append(text)

    # Assign Aadhaar number (take the first valid 12-digit number)
    if aadhaar_numbers:
        data["aadhaar_number"] = aadhaar_numbers[0]

    # Assign name (first English name found)
    if names_found:
        data["name"] = names_found[0].strip()

    # Build address
    if addresses:
        data["address"] = ", ".join(addresses)

    # Try to extract address from "Address:" label
    for line in lines:
        if 'address:' in line.lower():
            addr_part = re.split(r'[Aa]ddress\s*:', line, maxsplit=1)
            if len(addr_part) > 1 and addr_part[1].strip():
                data["address"] = addr_part[1].strip()
                break

    return data


# ── Excel Output ─────────────────────────────────────────────────────────────

DIRECTOR_COL = {1: 3, 2: 4}  # Column C=3, D=4


def populate_excel(pan_data: dict, aadhaar_data: dict, director_num: int,
                   template_path: str, output_path: str):
    """Fill the company formation Excel template with extracted data."""
    import openpyxl

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    col = DIRECTOR_COL[director_num]

    # PAN fields
    if pan_data:
        if pan_data.get("name"):
            ws.cell(row=4, column=col, value=pan_data["name"])
        if pan_data.get("fathers_name"):
            ws.cell(row=5, column=col, value=pan_data["fathers_name"])
        if pan_data.get("date_of_birth"):
            ws.cell(row=6, column=col, value=pan_data["date_of_birth"])
        if pan_data.get("pan_number"):
            ws.cell(row=16, column=col, value=pan_data["pan_number"])

    # Aadhaar fields
    if aadhaar_data:
        aadhaar_num = aadhaar_data.get("aadhaar_number", "")
        if aadhaar_num:
            ws.cell(row=17, column=col, value=f"AADHAR - {aadhaar_num}")
        if aadhaar_data.get("address"):
            ws.cell(row=18, column=col, value=aadhaar_data["address"])

    wb.save(output_path)
    print(f"\n✅ Excel saved: {output_path}")
    print(f"   Director {director_num} (Column {'C' if director_num == 1 else 'D'}) populated")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract PAN/Aadhaar data → Company Formation Excel"
    )
    parser.add_argument("--pan", type=str, help="Path to PAN card image")
    parser.add_argument("--aadhaar", type=str, help="Path to Aadhaar card image")
    parser.add_argument("--director", type=int, default=1, choices=[1, 2],
                        help="Director number (1 or 2, default: 1)")
    parser.add_argument("--template", type=str,
                        default="Checklist for Company Formation dated 11 02 2026.xlsx",
                        help="Path to Excel template")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel path")
    args = parser.parse_args()

    if not args.pan and not args.aadhaar:
        parser.error("Provide at least one of --pan or --aadhaar")

    for label, path in [("PAN", args.pan), ("Aadhaar", args.aadhaar)]:
        if path and not Path(path).exists():
            print(f"ERROR: {label} file not found: {path}")
            sys.exit(1)

    if not Path(args.template).exists():
        print(f"ERROR: Template not found: {args.template}")
        sys.exit(1)

    output_path = args.output or args.template.replace(".xlsx", "_filled.xlsx")

    # Initialize OCR once
    print("Initializing OCR engine...")
    t0 = time.time()
    ocr = get_ocr()
    print(f"OCR ready in {time.time() - t0:.1f}s")

    # Extract PAN
    pan_data = {}
    if args.pan:
        print(f"\n📄 Processing PAN card: {args.pan}")
        t0 = time.time()
        pan_lines = ocr_image(ocr, args.pan)
        print(f"   OCR lines ({time.time() - t0:.1f}s): {pan_lines}")
        pan_data = parse_pan(pan_lines)
        print(f"   Extracted: {json.dumps(pan_data, indent=2, ensure_ascii=False)}")

    # Extract Aadhaar
    aadhaar_data = {}
    if args.aadhaar:
        print(f"\n🆔 Processing Aadhaar card: {args.aadhaar}")
        t0 = time.time()
        aadhaar_lines = ocr_image(ocr, args.aadhaar)
        print(f"   OCR lines ({time.time() - t0:.1f}s): {aadhaar_lines}")
        aadhaar_data = parse_aadhaar(aadhaar_lines)
        print(f"   Extracted: {json.dumps(aadhaar_data, indent=2, ensure_ascii=False)}")

    # Write Excel
    populate_excel(pan_data, aadhaar_data, args.director, args.template, output_path)

    # Summary
    print("\n" + "=" * 60)
    print("EXTRACTION SUMMARY")
    print("=" * 60)
    fields_filled = 0
    if pan_data.get("name"):
        print(f"  Name:           {pan_data['name']}")
        fields_filled += 1
    if pan_data.get("fathers_name"):
        print(f"  Father's Name:  {pan_data['fathers_name']}")
        fields_filled += 1
    if pan_data.get("date_of_birth"):
        print(f"  DOB:            {pan_data['date_of_birth']}")
        fields_filled += 1
    if pan_data.get("pan_number"):
        print(f"  PAN:            {pan_data['pan_number']}")
        fields_filled += 1
    if aadhaar_data.get("aadhaar_number"):
        print(f"  Aadhaar No:     {aadhaar_data['aadhaar_number']}")
        fields_filled += 1
    if aadhaar_data.get("address"):
        print(f"  Address:        {aadhaar_data['address']}")
        fields_filled += 1
    print(f"\n  → {fields_filled}/6 fields populated in Excel")
    print(f"  → Output: {output_path}")


if __name__ == "__main__":
    main()
