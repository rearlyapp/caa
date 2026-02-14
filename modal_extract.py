#!/usr/bin/env python3
"""
Modal-deployed PAN + Aadhaar extraction with Qwen3-VL-2B on T4 GPU.

Usage:
    modal run modal_extract.py --pan pan.jpg --aadhaar aadhar.jpg --director 1
"""

import base64
import json
import re
import sys
import time
from pathlib import Path

import modal

# ── Modal App Setup ──────────────────────────────────────────────────────────

app = modal.App("caa-doc-extractor")

MODEL_ID = "Qwen/Qwen3-VL-2B-Instruct"
MODEL_DIR = "/model-cache"


def download_model():
    """Pre-download model weights into the image at build time."""
    from huggingface_hub import snapshot_download

    print(f"Downloading {MODEL_ID}...")
    snapshot_download(
        MODEL_ID,
        cache_dir=MODEL_DIR,
    )
    print(f"Model downloaded to {MODEL_DIR}")


image = (
    modal.Image.debian_slim(python_version="3.12")
    .pip_install(
        "torch",
        "torchvision",
        "transformers>=5.0.0",
        "accelerate",
        "Pillow",
        "openpyxl",
        "qwen-vl-utils",
    )
    .run_function(download_model)
)

# ── Prompts ──────────────────────────────────────────────────────────────────

PAN_PROMPT = """Look at this Indian PAN card image carefully.

On a PAN card, the layout is:
- Line 1: Cardholder's FULL NAME (e.g. "KISHORE SHEIK AHAMED M R")
- Line 2: Father's FULL NAME (e.g. "MOHAMMED RAHAMATHULLA")
- Line 3: Date of Birth (DD/MM/YYYY)
- Bottom: 10-character PAN number (e.g. BQJPK6347Q)

Extract these fields as JSON. The name and father's name are SEPARATE people on SEPARATE lines:

{
  "name": "cardholder name only (first name line)",
  "fathers_name": "father's name only (second name line)",
  "date_of_birth": "DD/MM/YYYY",
  "pan_number": "10 character PAN"
}

Return ONLY the JSON object, no other text."""

AADHAAR_PROMPT = """Look at this Aadhaar card image carefully.
Extract ALL the following fields and return ONLY valid JSON, nothing else:

{
  "name": "full name as printed",
  "aadhaar_number": "12 digit number (XXXX XXXX XXXX format)",
  "date_of_birth": "DD/MM/YYYY format",
  "gender": "MALE or FEMALE",
  "address": "complete residential address as printed including pincode"
}

Rules:
- Copy text EXACTLY as printed
- Aadhaar number is exactly 12 digits
- Include full address with pincode
- Return ONLY the JSON object, no other text"""


# ── Modal GPU Function ───────────────────────────────────────────────────────

@app.cls(
    image=image,
    gpu="T4",
    timeout=300,
    scaledown_window=120,
)
class Extractor:
    @modal.enter()
    def load_model(self):
        """Load model on container startup (cached across calls)."""
        import torch
        from transformers import Qwen3VLForConditionalGeneration, Qwen3VLProcessor

        print(f"Loading model: {MODEL_ID}")
        t0 = time.time()

        self.processor = Qwen3VLProcessor.from_pretrained(MODEL_ID, cache_dir=MODEL_DIR)
        self.model = Qwen3VLForConditionalGeneration.from_pretrained(
            MODEL_ID,
            dtype=torch.float16,
            cache_dir=MODEL_DIR,
            device_map="auto",
        )

        print(f"Model loaded in {time.time() - t0:.1f}s")

    @modal.method()
    def extract(self, image_b64: str, prompt: str) -> dict:
        """Extract fields from a base64-encoded image using the given prompt."""
        import torch
        from PIL import Image
        import io

        # Decode base64 image
        img_bytes = base64.b64decode(image_b64)
        image = Image.open(io.BytesIO(img_bytes)).convert("RGB")

        messages = [{
            "role": "user",
            "content": [
                {"type": "image", "image": image},
                {"type": "text", "text": prompt},
            ]
        }]

        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt",
        )
        inputs = inputs.to(self.model.device)

        t0 = time.time()
        with torch.no_grad():
            generated_ids = self.model.generate(**inputs, max_new_tokens=512)

        elapsed = time.time() - t0

        generated_ids_trimmed = [
            out_ids[len(in_ids):]
            for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        raw_text = self.processor.batch_decode(
            generated_ids_trimmed,
            skip_special_tokens=True,
            clean_up_tokenization_spaces=False,
        )[0]

        print(f"Inference time: {elapsed:.1f}s")
        print(f"Raw output:\n{raw_text}")

        return {"raw_text": raw_text, "elapsed": elapsed}


# ── JSON Parsing ─────────────────────────────────────────────────────────────

def parse_json_output(text: str) -> dict:
    """Extract JSON from model output."""
    # Try direct parse
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass

    # Try markdown code block
    match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    # Try first { ... } block
    match = re.search(r'\{[^{}]*\}', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass

    print("WARNING: Could not parse JSON from model output")
    return {}


# ── Excel Output ─────────────────────────────────────────────────────────────

DIRECTOR_COL = {1: 3, 2: 4}  # Column C=3, D=4


def populate_excel(pan_data: dict, aadhaar_data: dict, director_num: int,
                   template_path: str, output_path: str):
    """Fill the company formation Excel template with extracted data."""
    import openpyxl

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    col = DIRECTOR_COL[director_num]

    # PAN fields
    if pan_data:
        if pan_data.get("name"):
            ws.cell(row=4, column=col, value=pan_data["name"])
        if pan_data.get("fathers_name"):
            ws.cell(row=5, column=col, value=pan_data["fathers_name"])
        if pan_data.get("date_of_birth"):
            ws.cell(row=6, column=col, value=pan_data["date_of_birth"])
        if pan_data.get("pan_number"):
            ws.cell(row=16, column=col, value=pan_data["pan_number"])

    # Aadhaar fields
    if aadhaar_data:
        aadhaar_num = aadhaar_data.get("aadhaar_number", "")
        if aadhaar_num:
            ws.cell(row=17, column=col, value=f"AADHAR - {aadhaar_num}")
        if aadhaar_data.get("address"):
            ws.cell(row=18, column=col, value=aadhaar_data["address"])

    wb.save(output_path)
    print(f"\n✅ Excel saved: {output_path}")
    print(f"   Director {director_num} (Column {'C' if director_num == 1 else 'D'}) populated")


# ── Local Entrypoint ─────────────────────────────────────────────────────────

@app.local_entrypoint()
def main(
    pan: str = modal.parameter(default=""),
    aadhaar: str = modal.parameter(default=""),
    director: int = modal.parameter(default=1),
    output: str = modal.parameter(default=""),
):
    TEMPLATE = "Checklist for Company Formation dated 11 02 2026.xlsx"

    if not pan and not aadhaar:
        print("ERROR: Provide at least one of --pan or --aadhaar")
        sys.exit(1)

    for label, path in [("PAN", pan), ("Aadhaar", aadhaar)]:
        if path and not Path(path).exists():
            print(f"ERROR: {label} file not found: {path}")
            sys.exit(1)

    if not Path(TEMPLATE).exists():
        print(f"ERROR: Template not found: {TEMPLATE}")
        sys.exit(1)

    output_str = str(output).strip()
    output_path = output_str if (output_str and not output_str.startswith("<modal")) else TEMPLATE.replace(".xlsx", "_filled.xlsx")

    # Get remote extractor handle
    extractor = Extractor()

    total_t0 = time.time()

    # Extract PAN data
    pan_data = {}
    if pan:
        print(f"\n📄 Sending PAN card to GPU...")
        img_b64 = base64.b64encode(Path(pan).read_bytes()).decode()
        result = extractor.extract.remote(img_b64, PAN_PROMPT)
        print(f"   Inference: {result['elapsed']:.1f}s")
        print(f"   Raw: {result['raw_text']}")
        pan_data = parse_json_output(result["raw_text"])
        print(f"   Parsed: {json.dumps(pan_data, indent=2, ensure_ascii=False)}")

    # Extract Aadhaar data
    aadhaar_data = {}
    if aadhaar:
        print(f"\n🆔 Sending Aadhaar card to GPU...")
        img_b64 = base64.b64encode(Path(aadhaar).read_bytes()).decode()
        result = extractor.extract.remote(img_b64, AADHAAR_PROMPT)
        print(f"   Inference: {result['elapsed']:.1f}s")
        print(f"   Raw: {result['raw_text']}")
        aadhaar_data = parse_json_output(result["raw_text"])
        print(f"   Parsed: {json.dumps(aadhaar_data, indent=2, ensure_ascii=False)}")

    total_elapsed = time.time() - total_t0

    # Cross-validate names
    if pan_data and aadhaar_data:
        pan_name = pan_data.get("name", "").upper().strip()
        aad_name = aadhaar_data.get("name", "").upper().strip()
        if pan_name and aad_name and pan_name != aad_name:
            print(f"\n⚠️  Name mismatch:")
            print(f"   PAN:     {pan_data.get('name')}")
            print(f"   Aadhaar: {aadhaar_data.get('name')}")
            print(f"   Using PAN name (as per Excel template)")

    # Write Excel locally
    populate_excel(pan_data, aadhaar_data, director, TEMPLATE, output_path)

    # Summary
    print("\n" + "=" * 60)
    print("EXTRACTION SUMMARY")
    print("=" * 60)
    fields_filled = 0
    if pan_data.get("name"):
        print(f"  Name:           {pan_data['name']}")
        fields_filled += 1
    if pan_data.get("fathers_name"):
        print(f"  Father's Name:  {pan_data['fathers_name']}")
        fields_filled += 1
    if pan_data.get("date_of_birth"):
        print(f"  DOB:            {pan_data['date_of_birth']}")
        fields_filled += 1
    if pan_data.get("pan_number"):
        print(f"  PAN:            {pan_data['pan_number']}")
        fields_filled += 1
    if aadhaar_data.get("aadhaar_number"):
        print(f"  Aadhaar No:     {aadhaar_data['aadhaar_number']}")
        fields_filled += 1
    if aadhaar_data.get("address"):
        print(f"  Address:        {aadhaar_data['address']}")
        fields_filled += 1
    print(f"\n  ⏱️  Total time: {total_elapsed:.1f}s")
    print(f"  → {fields_filled}/6 fields populated in Director {director} column")
    print(f"  → Output: {output_path}")
